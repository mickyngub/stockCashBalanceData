import openpyxl
import datetime
from openpyxl.utils.cell import get_column_letter


workbook = openpyxl.load_workbook(filename="Main.xlsx")
sheet = workbook["CashBalanceJAN18"]
number_rows = sheet.max_row
number_columns = sheet.max_column
CURRENTDATE = "03/09/2021"
#######################

def findClosePrice():
        
    for i in range(1,number_rows):
        print("---------------finding ClosePrice----------------")
        try:
            tickerName = str(sheet[("A"+str(i+1))].value)
            endDateValue = str(sheet[("D"+str(i+1))].value)[:10]
            endDate = datetime.datetime.strptime(endDateValue, '%Y-%m-%d')
            # endDateM5 = endDate + datetime.timedelta(days=-5)

            strEndDate = endDate.strftime('%d/%m/%Y')
            # strEndDateM5 = endDateM5.strftime('%d/%m/%Y')
            currentDateOBJ = datetime.datetime.strptime(CURRENTDATE, "%d/%m/%Y")
            endDateOBJ = datetime.datetime.strptime(strEndDate, "%d/%m/%Y")
            print(tickerName + " cashBalance endDate " + strEndDate)
            
            tickerFileOpen = "cleanDate" + tickerName + ".xlsx"
            print(".....Opening '" + tickerFileOpen + "'....")

            tickerWorkBook = openpyxl.load_workbook(filename=tickerFileOpen)
            tickerSheet = tickerWorkBook.active
            tickerNumber_rows = tickerSheet.max_row
            # tickerNumber_columns = tickerSheet.max_column
            loopIndex = 0
            foundDate = False
            while foundDate == False:
                #if endDate is in the future, break the loop || loop more than 6 times 
                if foundDate == True or currentDateOBJ < endDateOBJ or loopIndex > 6:
                    break

                k = 0
                while k < tickerNumber_rows:
                    
                        
                    tickerDate = str(tickerSheet["A"+str(k+1)].value)
                    print('k value', k)
                    print('loopIndex value', loopIndex)
                    print('NORMAL ' + tickerDate)
                    strEndDateValue = strEndDate
                    if loopIndex > 0:
                        tempDate = datetime.datetime.strptime(strEndDate, '%d/%m/%Y') + datetime.timedelta(days=-loopIndex)
                        strEndDateValue = tempDate.strftime('%d/%m/%Y')

                        # strEndDate = str(tickerSheet["A"+str(k+1)].value)[:10]
                        # tickerDateDateTime = datetime.datetime.strptime(tickerDate, '%Y-%m-%d')
                        # tickerDateMinus1 = tickerDateDateTime + datetime.timedelta(days=-1)
                        # tickerDate = tickerDateMinus1.strftime('%d/%m/%Y')
                        print('MINUS ' + strEndDateValue)

                    #In case the endDate is on the holiday, the program cannot find the tickerDate which equals to strEndDate
                    
                    
                    # tickerCell = "A" + str(k+1)
                    # print('this is tickerdate', tickerDate, ' strEndDate', strEndDate)
                    if tickerDate == strEndDateValue:
                        foundDate = True

                        # print('Found EndDate at Cell Number'+tickerCell)
                        tickerDateM7 = str(tickerSheet["A"+str(k-6)].value)
                        tickerDateM6 = str(tickerSheet["A"+str(k-5)].value)
                        tickerDateM5 = str(tickerSheet["A"+str(k-4)].value)
                        tickerDateM4 = str(tickerSheet["A"+str(k-3)].value)
                        tickerDateM3 = str(tickerSheet["A"+str(k-2)].value)
                        tickerDateM2 = str(tickerSheet["A"+str(k-1)].value)
                        tickerDateM1 = str(tickerSheet["A"+str(k)].value)
                        tickerDate0 = str(tickerSheet["A"+str(k+1)].value)
                        tickerDateP1 = str(tickerSheet["A"+str(k+2)].value)
                        tickerDateP2 = str(tickerSheet["A"+str(k+3)].value)
                        tickerDateP3 = str(tickerSheet["A"+str(k+4)].value)
                        tickerDateP4 = str(tickerSheet["A"+str(k+5)].value)
                        tickerDateP5 = str(tickerSheet["A"+str(k+6)].value)

                        tickerCPEndDateM7 = str(tickerSheet["E"+str(k-6)].value)
                        tickerCPEndDateM6 = str(tickerSheet["E"+str(k-5)].value)
                        tickerCPEndDateM5 = str(tickerSheet["E"+str(k-4)].value)
                        tickerCPEndDateM4 = str(tickerSheet["E"+str(k-3)].value)
                        tickerCPEndDateM3 = str(tickerSheet["E"+str(k-2)].value)
                        tickerCPEndDateM2 = str(tickerSheet["E"+str(k-1)].value)
                        tickerCPEndDateM1 = str(tickerSheet["E"+str(k)].value)
                        tickerCPEndDate = str(tickerSheet["E"+str(k+1)].value)
                        tickerCPEndDateP1 = str(tickerSheet["E"+str(k+2)].value)
                        tickerCPEndDateP2 = str(tickerSheet["E"+str(k+3)].value)
                        tickerCPEndDateP3 = str(tickerSheet["E"+str(k+4)].value)
                        tickerCPEndDateP4 = str(tickerSheet["E"+str(k+5)].value)
                        tickerCPEndDateP5 = str(tickerSheet["E"+str(k+6)].value)          
                        
                        sheet["F"+str(i+1)] = float(tickerCPEndDateM7) if tickerCPEndDateM7 != "None" else "DNE"
                        sheet["G"+str(i+1)] = float(tickerCPEndDateM6) if tickerCPEndDateM6 != "None" else "DNE"
                        sheet["H"+str(i+1)] = float(tickerCPEndDateM5) if tickerCPEndDateM5 != "None" else "DNE"
                        sheet["I"+str(i+1)] = float(tickerCPEndDateM4) if tickerCPEndDateM4 != "None" else "DNE"
                        sheet["J"+str(i+1)] = float(tickerCPEndDateM3) if tickerCPEndDateM3 != "None" else "DNE"
                        sheet["K"+str(i+1)] = float(tickerCPEndDateM2) if tickerCPEndDateM2 != "None" else "DNE"     
                        sheet["L"+str(i+1)] = float(tickerCPEndDateM1) if tickerCPEndDateM1 != "None" else "DNE"
                        sheet["M"+str(i+1)] = float(tickerCPEndDate) if tickerCPEndDate != "None" else "DNE"
                        sheet["N"+str(i+1)] = float(tickerCPEndDateP1) if tickerCPEndDateP1 != "None" else "DNE"
                        sheet["O"+str(i+1)] = float(tickerCPEndDateP2) if tickerCPEndDateP2 != "None" else "DNE"
                        sheet["P"+str(i+1)] = float(tickerCPEndDateP3) if tickerCPEndDateP3 != "None" else "DNE"
                        sheet["Q"+str(i+1)] = float(tickerCPEndDateP4) if tickerCPEndDateP4 != "None" else "DNE"
                        sheet["R"+str(i+1)] = float(tickerCPEndDateP5) if tickerCPEndDateP5 != "None" else "DNE"
                        
                        print(tickerName + " " + tickerDateM7 + " " + tickerCPEndDateM7)
                        print(tickerName + " " + tickerDateM6 + " " + tickerCPEndDateM6)
                        print(tickerName + " " + tickerDateM5 + " " + tickerCPEndDateM5)
                        print(tickerName + " " + tickerDateM4 + " " + tickerCPEndDateM4)
                        print(tickerName + " " + tickerDateM3 + " " + tickerCPEndDateM3)
                        print(tickerName + " " + tickerDateM2 + " " + tickerCPEndDateM2)
                        print(tickerName + " " + tickerDateM1 + " " + tickerCPEndDateM1)
                        print("###")
                        print("END CashBalance " + tickerName + " " + tickerDate0 + " " + tickerCPEndDate)
                        print("###")
                        print(tickerName + " " + tickerDateP1 + " " + tickerCPEndDateP1)
                        print(tickerName + " " + tickerDateP2 + " " + tickerCPEndDateP2)
                        print(tickerName + " " + tickerDateP3 + " " + tickerCPEndDateP3)
                        print(tickerName + " " + tickerDateP4 + " " + tickerCPEndDateP4)
                        print(tickerName + " " + tickerDateP5 + " " + tickerCPEndDateP5)

                        workbook.save("CashBalanceCPJAN18.xlsx")
                        break
                    k += 1

                loopIndex += 1

                print("-------------------------------")

        except Exception as e: 
            print(e)
            print("Not a DATE")
            break

    workbook.save("CashBalanceCPJAN18.xlsx")

################

def findOpenPrice():

    for i in range(1, number_rows):
        print("---------------finding OpenPrice----------------")
        try:
            tickerName = str(sheet[("A"+str(i+1))].value)
            endDateValue = str(sheet[("D"+str(i+1))].value)[:10]
            endDate = datetime.datetime.strptime(endDateValue, '%Y-%m-%d')

            strEndDate = endDate.strftime('%d/%m/%Y')

            print(tickerName + " cashBalance endDate " + strEndDate)

            tickerFileOpen = "cleanDate" + tickerName + ".xlsx"
            print(".....Opening '" + tickerFileOpen + "'....")

            tickerWorkBook = openpyxl.load_workbook(filename=tickerFileOpen)
            tickerSheet = tickerWorkBook.active
            tickerNumber_rows = tickerSheet.max_row

            k = 0
            while k < tickerNumber_rows:
                tickerDate = str(tickerSheet["A"+str(k+1)].value)

                if tickerDate == strEndDate:
                    tickerDateM7 = str(tickerSheet["A"+str(k-6)].value)
                    tickerDateM6 = str(tickerSheet["A"+str(k-5)].value)
                    tickerDateM5 = str(tickerSheet["A"+str(k-4)].value)
                    tickerDateM4 = str(tickerSheet["A"+str(k-3)].value)
                    tickerDateM3 = str(tickerSheet["A"+str(k-2)].value)
                    tickerDateM2 = str(tickerSheet["A"+str(k-1)].value)
                    tickerDateM1 = str(tickerSheet["A"+str(k)].value)
                    tickerDate = str(tickerSheet["A"+str(k+1)].value)
                    tickerDateP1 = str(tickerSheet["A"+str(k+2)].value)
                    tickerDateP2 = str(tickerSheet["A"+str(k+3)].value)
                    tickerDateP3 = str(tickerSheet["A"+str(k+4)].value)
                    tickerDateP4 = str(tickerSheet["A"+str(k+5)].value)
                    tickerDateP5 = str(tickerSheet["A"+str(k+6)].value)

                    tickerOPEndDateM7 = str(tickerSheet["B"+str(k-6)].value)
                    tickerOPEndDateM6 = str(tickerSheet["B"+str(k-5)].value)
                    tickerOPEndDateM5 = str(tickerSheet["B"+str(k-4)].value)
                    tickerOPEndDateM4 = str(tickerSheet["B"+str(k-3)].value)
                    tickerOPEndDateM3 = str(tickerSheet["B"+str(k-2)].value)
                    tickerOPEndDateM2 = str(tickerSheet["B"+str(k-1)].value)
                    tickerOPEndDateM1 = str(tickerSheet["B"+str(k)].value)
                    tickerOPEndDate = str(tickerSheet["B"+str(k+1)].value)
                    tickerOPEndDateP1 = str(tickerSheet["B"+str(k+2)].value)
                    tickerOPEndDateP2 = str(tickerSheet["B"+str(k+3)].value)
                    tickerOPEndDateP3 = str(tickerSheet["B"+str(k+4)].value)
                    tickerOPEndDateP4 = str(tickerSheet["B"+str(k+5)].value)
                    tickerOPEndDateP5 = str(tickerSheet["B"+str(k+6)].value)

                    sheet["F"+str(i+1)] = float(tickerOPEndDateM7) if tickerOPEndDateM7 != "None" else "DNE"
                    sheet["G"+str(i+1)] = float(tickerOPEndDateM6) if tickerOPEndDateM6 != "None" else "DNE"
                    sheet["H"+str(i+1)] = float(tickerOPEndDateM5) if tickerOPEndDateM5 != "None" else "DNE"
                    sheet["I"+str(i+1)] = float(tickerOPEndDateM4) if tickerOPEndDateM4 != "None" else "DNE"
                    sheet["J"+str(i+1)] = float(tickerOPEndDateM3) if tickerOPEndDateM3 != "None" else "DNE"
                    sheet["K"+str(i+1)] = float(tickerOPEndDateM2) if tickerOPEndDateM2 != "None" else "DNE"     
                    sheet["L"+str(i+1)] = float(tickerOPEndDateM1) if tickerOPEndDateM1 != "None" else "DNE"
                    sheet["M"+str(i+1)] = float(tickerOPEndDate) if tickerOPEndDate != "None" else "DNE"
                    sheet["N"+str(i+1)] = float(tickerOPEndDateP1) if tickerOPEndDateP1 != "None" else "DNE"
                    sheet["O"+str(i+1)] = float(tickerOPEndDateP2) if tickerOPEndDateP2 != "None" else "DNE"
                    sheet["P"+str(i+1)] = float(tickerOPEndDateP3) if tickerOPEndDateP3 != "None" else "DNE"
                    sheet["Q"+str(i+1)] = float(tickerOPEndDateP4) if tickerOPEndDateP4 != "None" else "DNE"
                    sheet["R"+str(i+1)] = float(tickerOPEndDateP5) if tickerOPEndDateP5 != "None" else "DNE"
                  
                    print(tickerName + " " + tickerDateM7 + " " + tickerOPEndDateM7)
                    print(tickerName + " " + tickerDateM6 + " " + tickerOPEndDateM6)
                    print(tickerName + " " + tickerDateM5 + " " + tickerOPEndDateM5)
                    print(tickerName + " " + tickerDateM4 + " " + tickerOPEndDateM4)
                    print(tickerName + " " + tickerDateM3 + " " + tickerOPEndDateM3)
                    print(tickerName + " " + tickerDateM2 + " " + tickerOPEndDateM2)
                    print(tickerName + " " + tickerDateM1 + " " + tickerOPEndDateM1)
                    print("###")
                    print("END CashBalance " + tickerName + " " + tickerDate + " " + tickerOPEndDate)
                    print("###")
                    print(tickerName + " " + tickerDateP1 + " " + tickerOPEndDateP1)
                    print(tickerName + " " + tickerDateP2 + " " + tickerOPEndDateP2)
                    print(tickerName + " " + tickerDateP3 + " " + tickerOPEndDateP3)
                    print(tickerName + " " + tickerDateP4 + " " + tickerOPEndDateP4)
                    print(tickerName + " " + tickerDateP5 + " " + tickerOPEndDateP5)
                    workbook.save("CashBalanceOPJAN18.xlsx")

                    break
                k += 1
            print("-------------------------------")
        except:
            print("Not a DATE")
            break
    workbook.save("CashBalanceOPJAN18.xlsx")
    
userInput = input("OpenPrice / ClosePrice? ")
if userInput == "cp":
    findClosePrice()
elif userInput == "op":
    findOpenPrice()