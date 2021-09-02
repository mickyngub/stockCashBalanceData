import openpyxl
import csv
import datetime
from openpyxl.utils.cell import get_column_letter

tickerList = ["7UP", "ABM", "ADD", "AJA", "ARIN"]
replacement = "T02:00:00Z"

for k in tickerList:
    wb = openpyxl.Workbook()
    ws = wb.active
    csvFile = "SET_" + k + ", 1D.csv"
    with open(csvFile) as fileCSV:
        reader = csv.reader(fileCSV, delimiter=",")
        for row in reader:
            ws.append(row)
    wb.save("SET_" + k + ", 1D.xlsx")

for j in tickerList:
    xlsxFile = "SET_" + j + ", 1D.xlsx"
    workbook = openpyxl.load_workbook(filename=xlsxFile)
    sheet = workbook.active
    number_rows = sheet.max_row
    number_columns = 1


    for i in range(number_columns):
        for k in range(number_rows):
            cell = str(sheet[get_column_letter(i+1)+str(k+1)].value)
            # print(cell)
            if str(cell)[10:] == replacement:
                newCell = str(cell)[:10]
                date = datetime.datetime.strptime(newCell, '%Y-%m-%d').strftime('%d/%m/%Y')
                sheet[get_column_letter(i+1)+str(k+1)] = date
                
    fileSave = "cleanDate" + j + ".xlsx"
    workbook.save(fileSave)